import unittest
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains as ACT
from selenium.webdriver.chrome.options import Options as Opt
import openpyxl
import xlrd
from os.path import join, dirname, abspath
import logging
import unittest


'''# How to specify the desire location you want your downloaded files to be stored
chromeOptions = Opt()
chromeOptions.add_experimental_option('prefs', {'download.default_directory': 'C:/Users/Ntokozo/PycharmProjects'})
myBrowser = webdriver.Chrome(executable_path="C:/Users/Ntokozo/Downloads/chromedriver.exe",
                             chrome_options=chromeOptions)'''


myBrowser = webdriver.Chrome(executable_path="C:/Users/Ntokozo/Downloads/chromedriver.exe")
#myBrowser = webdriver.Firefox(executable_path="C:/Users/Ntokozo/Downloads/geckodriver.exe")

'''
fireFoxProf = webdriver.FirefoxProfile
fireFoxProf.set_preference('driver.helperApps.neverAsk.saveToDisk', 'text/plain, application/pdf')  # Mime type
fireFoxProf.set_preference('browser.download.manager.showWhenStarting', False)
fireFoxProf.set_preference('browser.download.dir', 'C:/Users/Ntokozo\PycharmProjects')
fireFoxProf.set_preference('browser.download.folderList', 2)
fireFoxProf.set_preference('pdf.js.disabled', True)
myBrowser = webdriver.Firefox(executable_path="C:/Users/Ntokozo/Downloads/geckodriver.exe", firefox_profile=fireFoxProf)
'''
'''
myBrowser.get("https://fred.stlouisfed.org/")
print("TITLE=", myBrowser.title)



elem = myBrowser.find_element_by_xpath('//*[@id="home-search"]/div/input')
elem.send_keys('gross gdp')
elem.send_keys(Keys.RETURN)
elem = myBrowser.find_element_by_xpath('//*[@id="search-results-list"]/li[1]/div/div[1]/a').click()
elem1 = myBrowser.find_element_by_xpath('//*[@id="input-cosd"]')
elem1.clear()
time.sleep(3)
elem1.send_keys('1960-10-1')

time.sleep(5)
myBrowser.get("https://www.49s.co.uk/49s/")
print('TITLE=',myBrowser.title)

# Go back and forward between two websites in a browser ...............................................................
time.sleep(4)
myBrowser.back()

time.sleep(4)
myBrowser.forward()
time.sleep(4)

myBrowser.quit()'''


# verify whether an input box is enabled or not..................................................................
myBrowser.get("https://www.linkedin.com/login?fromSignIn=true&trk=guest_homepage-basic_nav-header-signin")
print(myBrowser.title)


# WAITS are used to solve synchronization problems. we have Implicit and Explicit Waits..............................

myBrowser.implicitly_wait(5)  # wait for 5secs and this apply to elements
user_elem = myBrowser.find_element_by_name("session_key")
print(user_elem.is_displayed())  # returns boolean based on the result
print(user_elem.is_enabled())
print('User Entry cleared.', user_elem.clear())

password_elem = myBrowser.find_element_by_id("password")
print(password_elem.is_displayed())
print(password_elem.is_enabled())
print('Password Entry cleared.', password_elem.clear())
time.sleep(3)

user_elem.send_keys("aklamasyssy123@gmail.com")  # Sign-in
password_elem.send_keys("***********")
sign_in_btn = myBrowser.find_element_by_class_name("login__form_action_container ").click()    # Click a Button
time.sleep(5)

my_messages = myBrowser.find_element_by_id("messaging-nav-item").click()
time.sleep(5)

notification = myBrowser.find_element_by_id("notifications-nav-item").click()
time.sleep(5)
'''
'''
myBrowser.get("https://www.expedia.com/")
print(myBrowser.title)
myBrowser.implicitly_wait(2)
myBrowser.maximize_window()
flight_btn = myBrowser.find_element(By.XPATH, '//*[@id="tab-flight-tab-hp"]').click()
fly_from = myBrowser.find_element(By.ID, "flight-origin-hp-flight").send_keys("Johannesburg (JNB - O.R. Tambo Intl.)")
time.sleep(3)
destination = myBrowser.find_element(By.ID, "flight-destination-hp-flight").send_keys("New York (JFK - John F. Kennedy Intl.)")

depart_date = myBrowser.find_element(By.XPATH, "//*[@id='flight-departing-hp-flight']").clear()
depart_date = myBrowser.find_element(By.XPATH, "//*[@id='flight-departing-hp-flight']").send_keys("06/28/2020")

return_date = myBrowser.find_element(By.ID, 'flight-returning-hp-flight').clear()
return_date = myBrowser.find_element(By.ID, 'flight-returning-hp-flight').send_keys("07/28/2020")
search = myBrowser.find_element(By.XPATH, '//*[@id="gcw-flights-form-hp-flight"]/div[8]/label/button').click()


# Explicit wait statement
wait = WebDriverWait(myBrowser, 10)
nonStop = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="stopFilter_stops-2"]')))
nonStop.click()
myBrowser.close()



# Explicit wait statement
wait = WebDriverWait(myBrowser, 10)
nonStop = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="stopFilter_stops-2"]')))
nonStop.click()
#driver.close()



'''
# WORKING WITH INPUT BOXES..............how many input boxes, enter values and get status...........................
myBrowser.get("https://fs2.formsite.com/meherpavan/form2/index.html?1537702596407")
print(myBrowser.title)
textField = myBrowser.find_elements(By.CLASS_NAME, 'text_field')
print(len(textField))


firstName = myBrowser.find_element(By.NAME, 'RESULT_TextField-1').is_enabled()
firstName = myBrowser.find_element(By.NAME, 'RESULT_TextField-1').is_displayed()
firstName = myBrowser.find_element(By.NAME, 'RESULT_TextField-1').clear()
firstName = myBrowser.find_element(By.NAME, 'RESULT_TextField-1').send_keys('Sylvester')


lastName = myBrowser.find_element(By.NAME, 'RESULT_TextField-2').send_keys('Aklamavo')

firstName = myBrowser.find_element_by_name("RESULT_TextField-1")
print(firstName.is_displayed())  # returns boolean based on the result
print(firstName.is_enabled())
print('User Entry cleared.', firstName.clear())

# RADIO BUTTON , CHECKBOXES =====================================================================================

radiobtn = myBrowser.find_element(By.XPATH, '//*[@id="q26"]/table/tbody/tr[1]/td/label').is_selected()
print('radio Selected?',radiobtn)
radiobtn = myBrowser.find_element(By.XPATH, '//*[@id="q26"]/table/tbody/tr[1]/td/label').click()
radiobtn = myBrowser.find_element(By.XPATH, '//*[@id="q26"]/table/tbody/tr[1]/td/label').is_selected()
print('radio Selected?',radiobtn)

checkbox = myBrowser.find_element(By.XPATH, '//*[@id="q15"]/table/tbody/tr[1]/td/label').click()
checkbox = myBrowser.find_element(By.XPATH, '//*[@id="q15"]/table/tbody/tr[1]/td/label').is_selected()

# SELECT ITEMS FROM DROP-DOWN =======================================================================================

dropDown =myBrowser.find_element(By.CLASS_NAME, 'drop_down')
dropD = Select(dropDown)
#dropD.select_by_visible_text('Afternoon')
#dropD.select_by_value('Radio-0')
#dropD.select_by_index(2)
print(len(dropD.options))   # Know how many options available in the drop down
totalOptions = dropD.options
for option in totalOptions:
    print(option.text)      # we get the text values of all the available options.
'''

'''
# WORKING WITH LINKS ========================================================================================
myBrowser.get("http://sylvesteraklamavo.com/")
print(myBrowser.title)
blog = myBrowser.find_element(By.XPATH, '//*[@id="menu-item-389"]').click()
blog_links = myBrowser.find_elements(By.TAG_NAME, 'a')
print('Number of links:',len(blog_links))      # Number of links present on the web page.
for link in blog_links:
    print(link.text)        # print out the links

linkMetheny = myBrowser.find_element(By.LINK_TEXT, 'PAT METHENY ULTIMATE GUITAR LICKS').click()    # open a link
#linkMetheny = myBrowser.find_element(By.PARTIAL_LINK_TEXT, 'PAT').click()
myBrowser.maximize_window()
linkReverbnation = myBrowser.find_element(By.XPATH, '//*[@id="footer"]/div/div[1]/ul[2]/li[5]/a').click()
myBrowser.maximize_window()'''

'''
# WORKING WITH ALERTS ====================perform action on an element that can't be identified like pop-ups.===========
myBrowser.get("http://testautomationpractice.blogspot.com/")
print(myBrowser.title)
button = myBrowser.find_element(By.XPATH, '//*[@id="HTML9"]/div[1]/button').click()
myBrowser.switch_to.alert.accept()       # This will click the OK button on the alert/pop-up window.
#driver.switch_to.alert.dismiss()        # This will click the CANCEL button on the alert/pop-up window.

'''
'''
# WORKING WITH FRAMES =======================================================================================
myBrowser.get("https://www.selenium.dev/selenium/docs/api/java/index.html")
print(myBrowser.title)
time.sleep(3)
packageFrame = myBrowser.switch_to.frame('packageListFrame')    # this selects the selects frame 1, upper left.
link1 = myBrowser.find_element(By.LINK_TEXT, "org.openqa.selenium").click()

myBrowser.switch_to.default_content()           # takes back to initial display for you to select another frame.
time.sleep(3)
packageFrame = myBrowser.switch_to.frame('packageFrame')           # Frame 2
link2 = myBrowser.find_element(By.LINK_TEXT, "WebDriver.Window").click()

myBrowser.switch_to.default_content()
time.sleep(3)
packageFrame = myBrowser.switch_to.frame('classFrame')    # Frame 3
link3 = myBrowser.find_element(By.XPATH, '/html/body/div[4]/div[1]/ul/li/dl[1]/dd/a').click()
'''


'''# WORKING WITH BROWSER WINDOWS ==============================================================================
myBrowser.get("https://sylvesteraklamavo.com")
print(myBrowser.title)
myBrowser.find_element(By.XPATH, '//*[@id="footer"]/div/div[1]/ul[2]/li[5]/a').click()
time.sleep(2)
print(myBrowser.current_window_handle)  # returns value of the parent window (CDwindow-C54B8352BDD255BDD3439282A8D3148D)
tabs = myBrowser.window_handles
print(tabs) # returns values of all opened windows in the browser.
for tab in tabs:
    myBrowser.switch_to.window(tab)
    print(tab + ": " + myBrowser.title)  # get title of each opened tab.
    if myBrowser.title == "SYLVESTER AKLAMAVO | ReverbNation":
        time.sleep(3)
    else:
        print('Title Does Not Match')
        myBrowser.close()  # this closes the tab specified in the if statement. Which is the title of the child window
'''

'''
# WORKING WITH WEB TABLES IN BROWSER ===========================================NOT DONE YET==================================
myBrowser.get("https://www.investing.com/equities/bhs-tabletop-ag?cid=968851")
print(myBrowser.title)
rows = len(myBrowser.find_elements(By.XPATH, '//*[@id="leftColumn"]/table[1]/tbody/tr'))
#print(rows)
columns = len(myBrowser.find_elements(By.XPATH, '//*[@id="leftColumn"]/table[1]/thead/tr/th'))
#print(columns)
for row in range(1,rows):
    for col in range(1,columns):
        value = myBrowser.find_element_by_xpath('//*[@id="leftColumn"]/table["+str(row)+"]/tbody/tr["+str(col)+"]').text
        print(value)
'''


'''# WORKING ON SCROLLING ON A WEBPAGE ==================================================================================
myBrowser.get("https://www.countries-ofthe-world.com/flags-of-the-world.html")
print(myBrowser.title)
#browser.execute_script('window.scrollBy(0,3000)'" ")    # Scroll by pixel
benin = myBrowser.find_element(By.XPATH, '//*[@id="content"]/div[2]/div[2]/table[1]/tbody/tr[21]/td[1]/img')
myBrowser.execute_script("arguments[0].scrollIntoView();", benin)    # Scroll till element is found
myBrowser.execute_script('window.scrollBy(0,document.body.scrollHeight)')    # Scroll till the end of the page
'''

'''# MOUSE ACTIONS ==================================================================================
myBrowser.get("https://opensource-demo.orangehrmlive.com/")
print(myBrowser.title)
username = myBrowser.find_element(By.XPATH, '//*[@id="txtUsername"]').send_keys('Admin')
password = myBrowser.find_element(By.XPATH, '//*[@id="txtPassword"]').send_keys('admin123')
login = myBrowser.find_element(By.XPATH, '//*[@id="btnLogin"]').click()

admin = myBrowser.find_element(By.XPATH, '//*[@id="menu_admin_viewAdminModule"]')
userManagement = myBrowser.find_element(By.XPATH, '//*[@id="menu_admin_UserManagement"]')
user = myBrowser.find_element(By.XPATH, '//*[@id="menu_admin_viewSystemUsers"]')

actions = ACT(myBrowser)  # Mouse hover action using move_to_element
actions.move_to_element(admin).move_to_element(userManagement).move_to_element(user).click().perform()

myBrowser.get("http://testautomationpractice.blogspot.com/")
print(myBrowser.title)
myBrowser.maximize_window()
element = myBrowser.find_element(By.XPATH, '//*[@id="HTML10"]/div[1]/button')
action = ACT(myBrowser)
action.double_click(element).perform()  # Double click on element

myBrowser.get("https://swisnl.github.io/jQuery-contextMenu/demo.html")
print(myBrowser.title)
button = myBrowser.find_element(By.XPATH, '/html/body/div/section/div/div/div/p/span')
copy = myBrowser.find_element(By.XPATH,'/html/body/ul/li[3]/span')
action = ACT(myBrowser)
rightClick = action.context_click(button).perform()  # Right click on element
actionCopy = action.move_to_element(copy).click()

myBrowser.get("http://www.dhtmlgoodies.com/scripts/drag-drop-custom/demo-drag-drop-3.html")
print(myBrowser.title)
sourceElementRome = myBrowser.find_element(By.XPATH, '//*[@id="box6"]')
targetElementItaly = myBrowser.find_element(By.XPATH, '//*[@id="box106"]')
action = ACT(myBrowser)
action.drag_and_drop(sourceElementRome, targetElementItaly).perform()  # Drag & Drop action'''

'''
# UPLOAD A FILE =======================================================================================
myBrowser.get("http://testautomationpractice.blogspot.com/")
print(myBrowser.title)
myBrowser.maximize_window()
myBrowser.switch_to_frame(0)
item = '‪C:/Users/Ntokozo/Documents/BCS.pdf'
uploadButton = myBrowser.find_element(By.ID, 'RESULT_FileUpload-10')
uploadButton.send_keys(item)
action = ACT(myBrowser)
action.double_click(item).perform()  # Double click on element'''

'''# DOWNLOAD A FILE IN CHROME BROWSER=============================================================================
myBrowser.get("http://demo.automationtesting.in/FileDownload.html/")
print(myBrowser.title)
myBrowser.find_element(By.ID, 'textbox').send_keys('How are you today')
myBrowser.find_element(By.ID, 'createTxt').click()
myBrowser.find_element(By.ID, 'link-to-download').click()


# DOWNLOAD A FILE IN CHROME BROWSER=============================================================================
myBrowser.get("http://demo.automationtesting.in/FileDownload.html/")
print(myBrowser.title)
myBrowser.find_element(By.ID, 'textbox').send_keys('How are you today')
myBrowser.find_element(By.ID, 'createTxt').click()
myBrowser.find_element(By.ID, 'link-to-download').click()
'''

# DATA DRIVEN TESTING USING EXCEL ==================================================================================
'''

path = 'LNS14000024.xls'                   # Read data
workbook = xlrd.open_workbook(path)
sheet = workbook.sheet_by_index(0)
rows = sheet.cell(rowx=2,colx=6)
cols = sheet.max_column
print(rows)
print(cols)


path = "TestingSelenium.xlsx"               # Write data
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(row=r,column=c).value = 'Welcome'
        
workbook.save(path)


# OPERATION ON COOKIES =============================================================================================
myBrowser.get("https://www.49s.co.uk/49s/")
print(myBrowser.title)
cookies = myBrowser.get_cookies()
print(len(cookies))
print(cookies)
newCookie = {'name':'newcookie','value':'8888777'}
myBrowser.add_cookie(newCookie)
print(len(cookies))
print(cookies)
noCookies = myBrowser.delete_all_cookies()
print(noCookies)

# SCREENSHOT OPERATIONS =============================================================================================
myBrowser.get("http://sylvesteraklamavo.com/")
print(myBrowser.title)
myBrowser.maximize_window()
myBrowser.find_element_by_xpath('//*[@id="menu-item-387"]/a').click()
time.sleep(3)
myBrowser.save_screenshot('C:/Users/Ntokozo/Pictures/Saved Pictures1.jpg')

myBrowser.find_element_by_xpath('//*[@id="menu-item-391"]/a').click()
myBrowser.find_element_by_xpath('//*[@id="gal-container"]/ul/li[4]/a/img').click()
time.sleep(3)
myBrowser.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div/div[2]/a').click()
time.sleep(3)
myBrowser.get_screenshot_as_file('C:/Users/Ntokozo/Pictures/Saved Pictures2.png')


# LOG OPERATIONS =============================================================================================
# Log levels are: from Lowest to Highest - 1.DEBUG  2.INFO  3.WARNING  4.ERROR  5.CRITICAL
logging.basicConfig(filename='C:/Users/Ntokozo/PycharmProjects/SeleniumWebdriverTutorials/seleniumTestLog/Test.log',
                    format='%(asctime)s: %(levelname)s: %(message)s',
                    datefmt='%d/%m/%Y  %I:%M:%S %p',
                    level=logging.DEBUG,
                    )  # here we can use desire location
logging.debug("This is Debug Message")
logging.info("This is Info Message")
logging.warning("This is Warning Message")
logging.error("This is Error Message")
logging.critical("This is Critical Message")

# ALTERNATIVELY ======================== Set a logger object ====================================================
logging.basicConfig(filename='C:/Users/Ntokozo/PycharmProjects/SeleniumWebdriverTutorials/seleniumTestLog/Test2.log',
                    format='%(asctime)s: %(levelname)s: %(message)s',
                    datefmt='%d/%m/%Y  %I:%M:%S %p')

myLogger = logging.getLogger()
myLogger.setLevel(logging.DEBUG)

myLogger.debug("This is Debug Message")
myLogger.info("This is Info Message")
myLogger.warning("This is Warning Message")
myLogger.error("This is Error Message")
myLogger.critical("This is Critical Message")



# PYTHON UnitTesting FRAMEWORK ===================================================================================












myBrowser.get("https://data.sca.isr.umich.edu/data-archive/mine.php")
print(myBrowser.title)
time.sleep(3)


myBrowser.get("https://fred.stlouisfed.org/series/A191RL1Q225SBEA")
print(myBrowser.title)
time.sleep(3)



myBrowser.get("https://fred.stlouisfed.org/series/FEDFUNDS")
print(myBrowser.title)
time.sleep(3)
'''


